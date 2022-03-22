# For open the folders and files
import os
import pathlib

# For measuring the inference time and documentation of entry/exit time from the square??????
import time
from datetime import timedelta

folders_path = str(pathlib.Path(__file__).parent.resolve())  # נתיב תיקירת התוחנה
os.chdir(folders_path)

# For windows
import pygame
import PySimpleGUI as sg

pygame.mixer.init()

sg.theme('DarkAmber')  # שחור צהוב
layout_loading_window = [
    [sg.Image(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\loading.gif',
              key="Prog_bar")]]

loading_window = sg.Window('loading...', layout_loading_window)
for i in range(60):
    loading_window.read(timeout=40)  # 40ms per frame
    loading_window.Element("Prog_bar").UpdateAnimation(
        r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\loading.gif',
        time_between_frames=10)

import cv2
import easyocr

# For running inference on the TF-Hub module.
import tensorflow as tf
import tensorflow_hub as hub

# For drawing onto the image
import numpy as np
from PIL import Image
from PIL import ImageTk
from PIL import ImageDraw
from PIL import ImageFont
from PIL import ImageOps

# For analyze_range fun
import datetime
from datetime import date

import os

# for Timer Interrupt
import threading

# Yakir and Kfir Project Libraries
from Yakir_and_Kfir_Project import data_transfer
from Yakir_and_Kfir_Project import tables_fun

# For remove a folders
import shutil

# For data transfer library
import serial

# For administrator fun and Systems match file.xlsx
from openpyxl import load_workbook

# For System_boot fun and Systems match file.xlsx
import pandas as pd
from openpyxl import Workbook

ser = serial.Serial('COM6', '9600')
module_handle = "https://tfhub.dev/google/faster_rcnn/openimages_v4/inception_resnet_v2/1"
detector = hub.load(module_handle).signatures['default']
reader = easyocr.Reader(['en'])  # מודל לזיהוי מספרים ואותיות באנגלית!!!!
in_process = False
problems_dict = {}
statistics_time = 60.0

workbook = load_workbook(filename=r"venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\administrator.xlsx")
workbook.sheetnames
sheet = workbook.active
administrator_dict = {}
row = 1
while sheet["A{row}".format(row=row)].value is not None:
    administrator_dict[str(sheet["A{row}".format(row=row)].value)] = str(sheet["B{row}".format(row=row)].value)
    row += 1
channels = {}
number_of_lanes = 0
analyze_the_problems = False
analyze_time_frame = 1
clock_enabling = False
max_standby_time = []
time_document = time.strftime("%H_%M_%S")
problem_lane = '0'
checks_counter = 0
system_test = False
channel_test = False
min_cars_for_statistics = 5
min_len_level_3 = False
crowded_square = False
e_delay = []
time_crowded_square = timedelta(hours=int(time_document[0:2]), minutes=int(time_document[3:5]),
                                seconds=int(time_document[6:]))
problem_license_plate = ''

number_camera = 0
camera_connected = True
while camera_connected:
    globals()[f"camera{number_camera + 1}"] = cv2.VideoCapture(number_camera, cv2.CAP_DSHOW)
    camera_connected = globals()[f"camera{number_camera + 1}"].isOpened()
    if camera_connected:
        number_camera += 1

width = 1280
height = 856
list_object_names = [b'Truck', b'Bus', b'Van', b'Taxi', b'Ambulance', b'Car', b'Vehicle registration plate']

Level_2_length = 24  # The maximum number of cars in the traffic square + number_of_lanes
Level_3_length = 100  # The maximum number of cars in the exit lanes
level_1 = ['0', '0', '0', '0']
level_2 = {}
level_3 = {}

for i in range(60):
    loading_window.read(timeout=40)  # every 100ms, fire the event sg.TIMEOUT_KEY
    loading_window.Element("Prog_bar").UpdateAnimation(
        r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\loading.gif',
        time_between_frames=10)
loading_window.close()


# ========================================== fun main ======================================


def Camera_resolution(mode=''):
    if number_camera == 2:
        if mode == 'Systems matching':
            globals()[f"camera2"].set(3, 640)
            globals()[f"camera2"].set(4, 480)
        else:
            globals()[f"camera2"].set(3, 1280)
            globals()[f"camera2"].set(4, 720)


def resize_image(img_add, new_width=256, new_height=256):
    pil_image = Image.open(img_add)
    pil_image = ImageOps.fit(pil_image, (new_width, new_height), Image.ANTIALIAS)
    image_convert = pil_image.convert("RGB")
    image_convert.save(img_add, format="JPEG", quality=90)


def analyze_range():
    global analyze_the_problems
    global analyze_time_frame

    if analyze_the_problems:
        today = date.today()
        delta_time_days = datetime.timedelta(days=analyze_time_frame)
        ddr = str(today - delta_time_days)
        ddr = ddr[-2:] + '.' + ddr[5:7] + '.' + ddr[:4]
        if os.path.exists(r"Files\Problem report\{date}".format(date=ddr)):
            shutil.rmtree(r"Files\Problem report\{date}".format(date=ddr))
    else:
        today = date.today()
        delta_time_days = datetime.timedelta(days=1)
        ddr = str(today - delta_time_days)
        ddr = ddr[-2:] + '.' + ddr[5:7] + '.' + ddr[:4]
        if os.path.exists(r"Files\Problem report\{date}".format(date=ddr)):
            shutil.rmtree(r"Files\Problem report\{date}".format(date=ddr))


def image_capture(channel, mask_mode=False):
    global channels
    success_capture = False
    error_counter = -1
    while not success_capture:
        for i in range(2):
            success_capture, camera = globals()[f"camera{channels[channel][8]}"].read()
        error_counter += 1
        if error_counter > 3:
            print('Error:\n Camera{num} does not respond!!!'.format(num=channels[channel][8]))
            quit()
    if mask_mode:
        cv2.imwrite(r'Image mask.png', camera)
        resize_image(r'Image mask.png', width, height)
        if not (
        os.path.exists(r'C:\project\test\5555\venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Masks')):
            os.makedirs(r'C:\project\test\5555\venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Masks')
        cv2.imwrite(channels[channel][4], camera)
    else:
        cv2.imwrite(channels[channel][2], camera)


def enter_to_square(channel):
    time_now = time.strftime('%H:%M:%S')
    if channels[channel][6] != '':
        print(channel + ' enter to square')
        os.remove(channels[channel][5].format(image_name=channels[channel][6]))
        number_l = channels[channel][6]
        channels[channel][6] = ''
        level_2[number_l][0] = '1'
        level_2[number_l][4] = time_now


def draw_bounding_box_on_image(image, ymin, xmin, ymax, xmax, color, font=ImageFont.truetype("arial.ttf", 25),
                               thickness=4, display_str_list=()):
    """Adds a bounding box to an image"""
    draw = ImageDraw.Draw(image)
    im_width, im_height = image.size
    (left, right, top, bottom) = (xmin * im_width, xmax * im_width,
                                  ymin * im_height, ymax * im_height)
    draw.line([(left, top), (left, bottom), (right, bottom), (right, top),
               (left, top)],
              width=thickness,
              fill=color)

    display_str_heights = [font.getsize(ds)[1] for ds in display_str_list]
    total_display_str_height = (1 + 2 * 0.05) * sum(display_str_heights)

    if top > total_display_str_height:
        text_bottom = top
    else:
        text_bottom = top + total_display_str_height

    for display_str in display_str_list[::-1]:
        text_width, text_height = font.getsize(display_str)
        margin = np.ceil(0.05 * text_height)
        draw.rectangle([(left, text_bottom - text_height - 2 * margin),
                        (left + text_width, text_bottom)],
                       fill=color)
        draw.text((left + margin, text_bottom - text_height - margin),
                  display_str,
                  fill="black",
                  font=font)
        text_bottom -= text_height - 2 * margin
    return image


def draw_boxes_lane(image, result, max_boxes=100, min_score=0.1):
    boxes = result["detection_boxes"]
    class_names = result["detection_class_entities"]
    scores = result["detection_scores"]
    car_color = '#07b1fa'
    general_object_color = '#ffa703'
    cnt_cars = 0

    for i in range(min(boxes.shape[0], max_boxes)):
        if scores[i] >= min_score:
            if class_names[i].decode("ascii") == 'Car':
                cnt_cars += 1
                color = car_color
            else:
                color = general_object_color
            ymin, xmin, ymax, xmax = tuple(boxes[i])

            display_str = "{}: {}%: {}".format(class_names[i].decode("ascii"),
                                               int(100 * scores[i]), cnt_cars)
            image_pil = Image.fromarray(np.uint8(image)).convert("RGB")
            image_pil = draw_bounding_box_on_image(
                image_pil,
                ymin,
                xmin,
                ymax,
                xmax,
                color,
                display_str_list=[display_str])
            np.copyto(image, np.array(image_pil))
    return Image.fromarray(np.uint8(image)).convert("RGB")


def Resize_image_window(size, filename):  # size=(גובה ,רוחב)
    r_image = Image.open(filename)
    r_image = r_image.resize(size, resample=Image.BICUBIC)
    r_image = ImageTk.PhotoImage(image=r_image)
    return r_image


def Mask_on_image(channel):
    global width
    global height
    start_time = time.time()
    mask = Image.open(channels[channel][4])
    pixelMap = mask.load()
    image = Image.open(channels[channel][2])
    new_image = Image.new('RGB', (width, height))
    pixelsNew = new_image.load()
    pixelsOld = image.load()
    for i in range(mask.size[0]):
        for j in range(mask.size[1]):
            if pixelMap[i, j] == (0, 0, 0):
                pixelsNew[i, j] = (0, 0, 0, 255)
            else:
                pixelsNew[i, j] = pixelsOld[i, j]
    end_time = time.time()
    print("complete masking!")
    print("Mask preparation time:", end_time - start_time)
    new_image.save(channels[channel][2])
    mask.close()
    new_image.close()


def run_detector(detector, channel, draw_boxes=False):
    global reader
    global channels
    global channel_test

    min_score = 0.1
    cnt_cars = 0
    result = {}
    detection_boxes_list = []
    center_of_object_list = []
    detection_scores_list = []
    detection_class_entities_list = []
    if channels[channel][3] == 'on':
        Mask_on_image(channel)

    img = tf.io.read_file(channels[channel][2])
    img = tf.image.decode_jpeg(img, channels=3)
    converted_img = tf.image.convert_image_dtype(img, tf.float32)[tf.newaxis, ...]
    start_time = time.time()
    result1 = detector(converted_img)
    end_time = time.time()
    image_pil = Image.fromarray(np.uint8(img)).convert("RGB")
    im_width, im_height = image_pil.size

    result1 = {key: value.numpy() for key, value in result1.items()}

    for i in range(len(result1['detection_class_entities'])):
        if result1['detection_class_entities'][i] in list_object_names and result1['detection_scores'][i] >= min_score:
            detection_boxes_list += [list(result1['detection_boxes'][i])]
            ymin, xmin, ymax, xmax = tuple(result1['detection_boxes'][i])
            center_of_object_list += [[int(((xmin + xmax) * im_width) / 2), int(((ymin + ymax) * im_height) / 2)]]
            detection_scores_list += [result1['detection_scores'][i]]
            detection_class_entities_list += [result1['detection_class_entities'][i]]

    result['detection_boxes'] = np.array(detection_boxes_list)
    result['center_of_object'] = np.array(center_of_object_list)  # (x,y)
    result['detection_scores'] = np.array(detection_scores_list)
    result['detection_class_entities'] = np.array(detection_class_entities_list)

    print("Found %d objects." % len(result["detection_scores"]))
    print("Inference time: ", end_time - start_time)

    if channels[channel][0] == 'lane':
        for i in range(len(result["detection_scores"])):
            if result['detection_class_entities'][i] != list_object_names[-1]:
                cnt_cars += 1
        channels[channel][7] = str(cnt_cars)
        print("Number of cars identified is: %s" % channels[channel][7])
        if draw_boxes:
            image_with_bounding_box = draw_boxes_lane(img.numpy(), result)
            return image_with_bounding_box
        if not channel_test:
            level_1[channels[channel][1] - 1] = str(cnt_cars)
    else:
        license_plate_identified = False
        max_score = 0
        for i in range(len(result["detection_scores"])):
            if result['detection_class_entities'][i] == list_object_names[-1] and \
                    result['detection_scores'][i] >= max_score:
                license_plate_identified = True
                max_score = result['detection_scores'][i]
                ymin, xmin, ymax, xmax = tuple(result['detection_boxes'][i])

        if license_plate_identified:
            image = img.numpy()
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            x1, x2, y1, y2 = int(xmin * im_width) + 3, int(xmax * im_width) + 3, int(ymin * im_height) + 3, int(
                ymax * im_height) + 3
            cropped_image = gray[y1:y2, x1:x2]
            result2 = reader.readtext(cropped_image)
            print(result2)

            for i in result2:
                number_l = ''
                for c in i[-2]:
                    if c in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9']:
                        number_l += c
                    elif c.isalnum():
                        number_l = 'xxxxxxxxx'
                        break
                if not 9 > len(number_l) > 5:
                    number_l = 'xxxxxxxxx'
                if number_l != 'xxxxxxxxx':
                    print(number_l)
                    if channels[channel][0] == 'in':
                        if number_l in level_2.keys() and level_2[number_l][0] == '1':
                            del level_2[number_l]
                        elif number_l in level_2.keys() and level_2[number_l][0] == '0':
                            os.remove(channels[f'channel_{level_2[number_l][2]}_in'][5].format(
                                image_name=channels[f'channel_{level_2[number_l][2]}_in'][6]))
                            channels[f'channel_{level_2[number_l][2]}_in'][6] = ''
                        if number_l in level_3.keys():
                            del level_3[number_l]
                        channels[channel][6] = number_l
                        if not channel_test:
                            time_now = time.strftime('%M:%S')
                            level_2[number_l] = ['0', '00:00', str(channels[channel][1]), '0', '00:00:00', '00:00:00',
                                                 time_now]
                            license_plates_image = Image.fromarray(np.uint8(cropped_image)).convert("RGB")
                            license_plates_image.save(channels[channel][5].format(image_name=number_l))
                    else:
                        if number_l in level_2:
                            time_now = time.strftime('%H:%M:%S')
                            level_3[number_l] = level_2[number_l]
                            if level_2[number_l][0] == '0':
                                enter_to_square(f'channel_{level_2[number_l][2]}_in')
                            del level_2[number_l]
                            level_3[number_l][3] = str(channels[channel][1])
                            level_3[number_l][5] = time_now
                            level_3[number_l][0] = '2'
                            print("Vehicle number %s left the square" % number_l)
                    break
                else:
                    print(number_l)

            if draw_boxes:
                image_pil = Image.fromarray(np.uint8(image)).convert("RGB")
                image_with_bounding_box = \
                    draw_bounding_box_on_image(image_pil, ymin, xmin, ymax, xmax, '#f20707',
                                               font=ImageFont.truetype("arial.ttf", 65), thickness=4,
                                               display_str_list=[number_l])
                return image_with_bounding_box
        else:
            img = Image.open(channels[channel][2])
            return img


def detect_img(channel, draw_boxes=False, without_camera=False):
    if not without_camera:
        image_capture(channel)
    if not os.path.exists(channels[channel][2]):
        print('No image found!!!')
        return None
    start_time = time.time()
    resize_image(channels[channel][2], width, height)
    image_with_bounding_box = run_detector(detector, channel, draw_boxes)
    end_time = time.time()
    print("Total time: ", end_time - start_time)
    return image_with_bounding_box


def Check_transmission():
    global system_test
    global time_document
    global problem_lane
    global checks_counter
    global crowded_square
    global level_1, level_2, level_3

    vehicles_number_in_lane = 0
    vehicles_number_enter = 0

    if not system_test:
        detect_img(f'channel_{problem_lane}_lane')
        print('problem lane = ', problem_lane)
    vehicles_number_in_lane = int(level_1[problem_lane - 1])

    time_document_convert = timedelta(hours=int(time_document[0:2]),
                                      minutes=int(time_document[3:5]),
                                      seconds=int(time_document[6:]))

    for number_l in level_2.keys():
        level_2_convert = timedelta(hours=int(level_2[number_l][4][0:2]),
                                    minutes=int(level_2[number_l][4][3:5]),
                                    seconds=int(level_2[number_l][4][6:]))
        if level_2[number_l][0] != '0' and int(level_2[number_l][2]) == problem_lane and \
                level_2_convert > time_document_convert:
            vehicles_number_enter += 1
    for number_l in level_3.keys():
        level_3_convert = timedelta(hours=int(level_3[number_l][4][0:2]),
                                    minutes=int(level_3[number_l][4][3:5]),
                                    seconds=int(level_3[number_l][4][6:]))
        if level_3[number_l][0] != '0' and int(level_3[number_l][2]) == problem_lane and \
                level_3_convert > time_document_convert:
            vehicles_number_enter += 1
    if crowded_square:
        vehicles_number_in_lane = 0
        vehicles_number_enter = 0
    check = []
    check += [vehicles_number_in_lane]
    check += [vehicles_number_enter]
    print(f'CHECK {checks_counter + 1} = ', check)
    tables_fun.creating_problem_document_CHECK(check, checks_counter + 1,
                                               time_document, system_test)
    data_transfer.table_transmission(ser, 'Check',
                                     vehicles_number_in_lane=vehicles_number_in_lane,
                                     vehicles_number_enter=vehicles_number_enter)


def Int_time():
    global system_test
    global clock_enabling
    global problems_dict
    global in_process
    global max_standby_time
    global channels
    global time_document
    global problem_license_plate
    global problem_lane
    global min_len_level_3
    global statistics_time
    global crowded_square
    global time_crowded_square
    global level_1, level_2, level_3
    global min_cars_for_statistics
    global e_delay

    if clock_enabling:
        time_now = time.strftime('%H:%M:%S')
        time_now_convert = timedelta(hours=int(time_now[0:2]), minutes=int(time_now[3:5]),
                                     seconds=int(time_now[6:]))
        standing_cars = 0
        for number_l in level_2.keys():
            if level_2[number_l][0] == '1':
                time_in_convert = timedelta(hours=int(level_2[number_l][4][0:2]),
                                            minutes=int(level_2[number_l][4][3:5]),
                                            seconds=int(level_2[number_l][4][6:]))
                if (time_now_convert - time_in_convert) > timedelta(minutes=(statistics_time / 2)):
                    standing_cars += 1
        if standing_cars > 2 and not system_test:
            print('Crowded square!!!')
            crowded_square = True
            level_1 = ['0', '0', '0', '0']
            level_2 = {}
            level_3 = {}
            print('level_1:')
            print(level_1)
            print('level_2:')
            print(level_2)
            print('level_3:')
            print(level_3)
            time_crowded_square = time_now_convert
        if crowded_square and (time_now_convert - time_crowded_square) > timedelta(minutes=statistics_time):
            crowded_square = False
            print('Active square!!!')

        analyze_range()
        if not system_test:
            del_level_2_ket = []
            for number_l in level_2.keys():
                if level_2[number_l][0] == '1':
                    time_in_convert = timedelta(hours=int(level_2[number_l][4][0:2]),
                                                minutes=int(level_2[number_l][4][3:5]),
                                                seconds=int(level_2[number_l][4][6:]))
                    if (time_now_convert - time_in_convert) > timedelta(minutes=statistics_time):
                        del_level_2_ket += [number_l]
            if len(del_level_2_ket) > 0:
                print('remove cars from Level 2 Table because statistics time has passed:')
                for j in range(len(del_level_2_ket)):
                    print(del_level_2_ket[j])
                for number_l in del_level_2_ket:
                    del level_2[number_l]
            if len(level_2) > Level_2_length:
                for i in range(len(level_2) - Level_2_length):
                    del_vector = ''
                    delta_time_del_vector = timedelta(seconds=0)
                    for number_l in level_2.keys():
                        time_in_convert = timedelta(hours=int(level_2[number_l][4][0:2]),
                                                    minutes=int(level_2[number_l][4][3:5]),
                                                    seconds=int(level_2[number_l][4][6:]))
                        if (time_now_convert - time_in_convert) > delta_time_del_vector:
                            delta_time_del_vector = time_now_convert - time_in_convert
                            del_vector = number_l
                    del level_2[del_vector]

            del_level_3_ket = []
            for number_l in level_3.keys():
                time_out_convert = timedelta(hours=int(level_3[number_l][5][0:2]),
                                             minutes=int(level_3[number_l][5][3:5]),
                                             seconds=int(level_3[number_l][5][6:]))
                if (time_now_convert - time_out_convert) > timedelta(minutes=statistics_time):
                    del_level_3_ket += [number_l]
            if len(del_level_3_ket) > 0:
                print('remove cars from Level 3 Table because statistics time has passed:')
                for j in range(len(del_level_3_ket)):
                    print(del_level_3_ket[j])
                for number_l in del_level_3_ket:
                    del level_3[number_l]
            if len(level_3) > Level_3_length:
                for i in range(len(level_3) - Level_3_length):
                    del_vector = ''
                    delta_time_del_vector = timedelta(seconds=0)
                    for number_l in level_3.keys():
                        time_out_convert = timedelta(hours=int(level_3[number_l][5][0:2]),
                                                     minutes=int(level_3[number_l][5][3:5]),
                                                     seconds=int(level_3[number_l][5][6:]))
                        if (time_now_convert - time_out_convert) > delta_time_del_vector:
                            delta_time_del_vector = time_now_convert - time_out_convert
                            del_vector = number_l
                    del level_3[del_vector]
        if len(level_3) < min_cars_for_statistics:
            min_len_level_3 = False
        else:
            min_len_level_3 = True
        
        for number_l in level_2.keys():
            if level_2[number_l][0] == '0':
                time_now_2 = time.strftime('%M:%S')
                time_now_convert_2 = timedelta(minutes=int(time_now_2[0:2]), seconds=int(time_now_2[3:]))
                time_old_convert = timedelta(minutes=int(level_2[number_l][6][0:2]),
                                             seconds=int(level_2[number_l][6][3:]))
                standby_time = time_now_convert_2 - time_old_convert
                level_2[number_l][1] = str(standby_time)[2:]    
        
        if not in_process:
            for i in range(number_of_lanes):
                if not channels[f'channel_{i + 1}_in'][9]:
                    time_now = time.strftime('%H:%M:%S')
                    time_now_convert = timedelta(hours=int(time_now[0:2]), minutes=int(time_now[3:5]),
                                                 seconds=int(time_now[6:]))
                    interruption_lock_time = channels[f'channel_{i + 1}_in'][10]
                    interruption_lock_time_convert = timedelta(hours=int(interruption_lock_time[0:2]),
                                                               minutes=int(interruption_lock_time[3:5]),
                                                               seconds=int(interruption_lock_time[6:]))
                    if (time_now_convert - interruption_lock_time_convert) > timedelta(seconds=e_delay[i]):
                        channels[f'channel_{i + 1}_in'][9] = True

            for number_l in level_2.keys():
                if level_2[number_l][0] == '0':
                    time_now_2 = time.strftime('%M:%S')
                    time_now_convert_2 = timedelta(minutes=int(time_now_2[0:2]), seconds=int(time_now_2[3:]))
                    time_old_convert = timedelta(minutes=int(level_2[number_l][6][0:2]),
                                                 seconds=int(level_2[number_l][6][3:]))
                    standby_time = time_now_convert_2 - time_old_convert
                    level_2[number_l][1] = str(standby_time)[2:]
                    if len(max_standby_time) == 0:
                        print('Error!')
                    elif channels['channel_{i}_in'.format(i=level_2[number_l][2])][9] and min_len_level_3:
                        if standby_time > max_standby_time[int(level_2[number_l][2]) - 1]:
                            print(number_l)
                            problems_dict[number_l] = time_now_convert_2
            if len(problems_dict) > 0:
                problem_lane = '0'
                max_standby = timedelta(minutes=0, seconds=0)
                problem_license_plate = ''
                for number_l in problems_dict.keys():
                    if problems_dict[number_l] > max_standby:
                        problem_license_plate = number_l
                        max_standby = problems_dict[number_l]
                        problem_lane = int(level_2[number_l][2])
                if problem_lane > 0:
                    print('problem_lane = ', problem_lane)
                    in_process = True
                    if not system_test:
                        for number in range(number_of_lanes):
                            print(f'channel_{number + 1}_lane')
                            detect_img(f'channel_{number+1}_lane')

                    channels['channel_{i}_in'.format(i=level_2[number_l][2])][10] = time.strftime('%H:%M:%S')
                    channels['channel_{i}_in'.format(i=level_2[number_l][2])][9] = False
                    time_document = time.strftime("%H_%M_%S")
                    tables_fun.creating_problem_document_start(level_1, level_2, level_3, problem_license_plate,
                                                               time_document, system_test)
                    data_transfer.table_transmission(ser, 'Tables', level_1=level_1, level_2=level_2,
                                                     level_3=level_3, problem_lane=problem_lane)
                    problems_dict = {}
        interrupt_time = threading.Timer(1.0, Int_time)
        interrupt_time.start()


def Building_channel_dictionary():
    global channels
    global number_of_lanes
    global max_standby_time

    match_file = pd.read_excel(
        r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Systems match file.xlsx')
    lane_file = pd.read_excel(
        r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Lane map.xlsx')
    max_standby_time = []
    for i in range(number_of_lanes):
        max_standby_time += [timedelta(seconds=int(lane_file['standby'][i]))]
    camera_index = 0
    for i in range(number_of_lanes):
        for x in ['in', 'out', 'lane']:
            channels['channel_{i}_{x}'.format(i=i + 1, x=x)] = \
                [x, i + 1, r'channels\channel {i}\{x}\camera.jpg'.format(i=i + 1, x=x), 'off',
                 r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Masks\mask channel {i} {x}.png'.format(
                     i=i + 1, x=x),
                 '', '', 0, match_file['Camera'][camera_index], True, '', False, 0]
            if os.path.exists(
                    r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Masks\mask channel {i} {x}.png'.format(
                            i=i + 1, x=x)):
                channels['channel_{i}_{x}'.format(i=i + 1, x=x)][3] = 'on'
            if x == 'in':
                channels['channel_{i}_{x}'.format(i=i + 1, x=x)][5] = \
                    r'channels\channel {i}\in'.format(i=i + 1) + r'\License plate images\{image_name}.jpg'
                channels['channel_{i}_{x}'.format(i=i + 1, x=x)][10] = time.strftime('%H:%M:%S')
            camera_index += 1

    """
    channels:
    *type: dictionary
    *Data Structure:
        * key = "channel_num_class" (num = 1,2,...,N ; class = in, out, lane)
        * value : type = list 
            [0] Channel class = in, out, lane : type = str
            [1] Channel number = 1,2,...,N : type = int
            [2] Camera image address : type = str
            [3] Mask mode = on, off : type = str
            [4] Mask address : type = str
            [5] License plate image address (if class = in) : type = str
            [6] License plate number (if class = in/out) : type = str
            [7] Number of cars in the lane (if class = lane) : type = int
            [8] Number camera : type = numpy.int64
            [9] Enabling channel interference (if class = in) type = bool
            [10] Interruption lock time (00:00:00) (if class = in) : type = str
            [11] Image capture (if class = out) : type = bool
            [12] Counter failures = 0, 1, 2 (if class = out) : type = int
    """
    list_names_folders = ['channels', r'\channel {num_channel}', r'\in\License plate images', r'\out', r'\lane']
    if not (os.path.exists(list_names_folders[0])):
        os.makedirs(list_names_folders[0])
        for i in range(4):
            os.makedirs(list_names_folders[0] + list_names_folders[1].format(num_channel=str(i + 1)))
            os.makedirs(
                list_names_folders[0] + list_names_folders[1].format(num_channel=str(i + 1)) + list_names_folders[2])
            os.makedirs(
                list_names_folders[0] + list_names_folders[1].format(num_channel=str(i + 1)) + list_names_folders[3])
            os.makedirs(
                list_names_folders[0] + list_names_folders[1].format(num_channel=str(i + 1)) + list_names_folders[4])
    else:
        for i in range(4):
            shutil.rmtree(r'channels\channel {num_channel}\in\License plate images'.format(num_channel=str(i + 1)))
            os.makedirs(r'channels\channel {num_channel}\in\License plate images'.format(num_channel=str(i + 1)))


def Lane_map_transfer():
    lane_map = pd.read_excel(
        r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Lane map.xlsx')
    time_in = lane_map['Time in'].tolist()
    cars_in = lane_map['Cars in'].tolist()
    data_transfer.table_transmission(ser, 'Lane map', time_in=time_in, cars_in=cars_in)


def checking_core_files():
    list_files = []
    missing_files = 0
    if not os.path.exists(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Boot file.xlsx'):
        missing_files += 1
        list_files += ['Boot file']
    if not os.path.exists(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Systems match file.xlsx'):
        missing_files += 1
        list_files += ['Systems matching']
    if not os.path.exists(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Lane map.xlsx'):
        missing_files += 1
        list_files += ['Lane map']
    return missing_files > 0, list_files


def Update_boot_file_parameters_and_Lane_map_transfer():
    global number_of_lanes
    global analyze_the_problems
    global analyze_time_frame
    global Level_2_length
    global Level_3_length
    global min_cars_for_statistics
    global statistics_time
    global e_delay

    check, files = checking_core_files()
    if 'Boot file' not in files:
        boot_file = pd.read_excel(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Boot file.xlsx')
        analyze_the_problems = bool(boot_file['analyze'].tolist()[-1])
        analyze_time_frame = boot_file['days'].tolist()[-1]
        number_of_lanes = boot_file['lanes'].tolist()[-1]
        Level_2_length = boot_file['Level_2'].tolist()[-1]
        Level_3_length = boot_file['Level_3'].tolist()[-1]
        min_cars_for_statistics = boot_file['min cars'].tolist()[-1]
        statistics_time = boot_file['statistics'].tolist()[-1]

    if 'Lane map' not in files:
        lane_map = pd.read_excel(
            r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Lane map.xlsx')
        e_delay = lane_map['E delay'].tolist()
        Lane_map_transfer()

    if not check:
        Building_channel_dictionary()


# ========================================== win administrator mode ======================================
def goodbye():
    layout_windows_goodbye = [
        [sg.Image(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\goodbye.png')]]
    pygame.mixer.music.load(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\sounds\goodbye.wav')
    pygame.mixer.music.play(loops=0)
    goodbye_window = sg.Window("GoodBye", layout_windows_goodbye)
    goodbye_window.read(timeout=2000)
    goodbye_window.close()
    time.sleep(5)
    quit()


def System_test():
    global clock_enabling
    global in_process
    global system_test
    global level_1
    global level_2
    global level_3
    global time_document
    global problem_license_plate
    global checks_counter
    check, files = checking_core_files()
    if check:
        layout_missing_core_files = [
            [sg.Image(
                r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\Exclamation mark.png'),
                sg.Text('Missing core files!\nYou must first finish setting up the core files!')]]
        missing_core_files_window = sg.Window('Error', layout_missing_core_files)
        pygame.mixer.music.load(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\sounds\error.wav')
        pygame.mixer.music.play(loops=0)
        missing_core_files_window.read()
        missing_core_files_window.close()
        return
    Update_boot_file_parameters_and_Lane_map_transfer()
    system_test = True
    if not (os.path.exists(r"Files\System test")):
        os.makedirs(r"Files\System test")
    if not (os.path.exists(r"Files\System test\manual file.xlsx")):
        tables_fun.manual_file(number_of_lanes)
    while True:
        checks = []
        checks_counter = 0
        layout_windows_System_test = [
            [sg.Text(size=2),
             sg.Image(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\logo.png')],
            [sg.Text(font=("Arial", "2"))],
            [sg.Text('Please build the "manual file" and "checks" then click run', font=("Arial", "12"), size=(36, 2))],
            [sg.Text()],
            [sg.Text(size=12), sg.Text('Cars waiting'), sg.Text(size=5), sg.Text('Cars in')],
            [sg.Text(size=1), sg.Text('CHECK1'), sg.Text(size=4),
             sg.Spin([i for i in range(0, 100)], size=2, key='wait1'),
             sg.Text(size=9), sg.Spin([i for i in range(0, 100)], size=2, key='in1')],
            [sg.Checkbox('CHECK2', key='CHECK2'), sg.Text(size=4),
             sg.Spin([i for i in range(0, 100)], size=2, key='wait2'),
             sg.Text(size=9), sg.Spin([i for i in range(0, 100)], size=2, key='in2')],
            [sg.Checkbox('CHECK3', key='CHECK3'), sg.Text(size=4),
             sg.Spin([i for i in range(0, 100)], size=2, key='wait3'),
             sg.Text(size=9), sg.Spin([i for i in range(0, 100)], size=2, key='in3')],
            [sg.Text(font=("Arial", "2"))],
            [sg.Text(font=("Arial", "25"), size=6), sg.Button('Run')]]
        System_test_window = sg.Window("System test", layout_windows_System_test)
        event, values = System_test_window.read()
        if event == 'Run':
            checks += [[values['wait1'], values['in1']]]
            if values['CHECK2']:
                checks += [[values['wait2'], values['in2']]]
            if values['CHECK3']:
                checks += [[values['wait3'], values['in3']]]
            if len(checks) < 3:
                while len(checks) < 3:
                    checks += [checks[-1]]
            print(checks)

            System_test_window.close()
            level_1, level_2, level_3 = tables_fun.manual_operation(number_of_lanes)
            if len(level_1) == 0 and len(level_2) == 0 and len(level_3) == 0:
                print('''Please check the "manual file"\nThe "manual file" is not found or is invalid!''')
            else:
                print("System test running")
                print('level_1')
                print(level_1)
                print('level_2:')
                for v in level_2.items():
                    print("%s -> %s" % v)
                print('level_3:')
                for v in level_3.items():
                    print("%s -> %s" % v)
                clock_enabling = True
                interrupt_time = threading.Timer(1.0, Int_time)
                interrupt_time.start()
                layout_windows_System_test_run = [
                    [sg.Image(
                        r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\logo.png')],
                    [sg.Text(font=("Arial", "2"))],
                    [sg.Text('System test running!', font=("Arial", "12"), size=(20, 2))]]
                System_test_run_window = sg.Window("System test running", layout_windows_System_test_run)
                while True:
                    event_System_test_run, values = System_test_run_window.read(timeout=40)
                    if event_System_test_run == sg.WIN_CLOSED:
                        clock_enabling = False
                        in_process = False
                        level_1 = ['0', '0', '0', '0']
                        level_2 = {}
                        level_3 = {}
                        System_test_run_window.close()
                        break
                    data = data_transfer.read_command(ser)
                    if data != '':
                        if 'ut ' in data:
                            tables_fun.creating_problem_document_UT(data, time_document, system_test)
                        elif 'pt ' in data:
                            tables_fun.creating_problem_document_PT(data, time_document, system_test)
                        elif data == 'CHECK':
                            print(f'CHECK {checks_counter + 1} = ', checks[checks_counter])
                            tables_fun.creating_problem_document_CHECK(checks[checks_counter], checks_counter + 1,
                                                                       time_document, system_test)
                            vehicles_number_in_lane = checks[checks_counter][0]
                            vehicles_number_enter = checks[checks_counter][1]
                            data_transfer.table_transmission(ser, 'Check',
                                                             vehicles_number_in_lane=vehicles_number_in_lane,
                                                             vehicles_number_enter=vehicles_number_enter)
                            checks_counter += 1
                        elif 'RT ' in data:
                            tables_fun.creating_problem_document_RT(data, time_document, system_test)
                            print([data])
                        elif 'SOLVE = ' in data:
                            tables_fun.creating_problem_document_SOLVE(data, time_document, system_test)
                            tables_fun.creating_problem_document_end(level_1, level_2, level_3, problem_license_plate,
                                                                     time_document, system_test)
                            print([data])
                            in_process = False
                            clock_enabling = False
                            level_1 = ['0', '0', '0', '0']
                            level_2 = {}
                            level_3 = {}
                            System_test_run_window.close()
        if event == sg.WIN_CLOSED:
            System_test_window.close()
            break
    system_test = False


def Building_mask():
    global width
    global height
    global channels
    global number_of_lanes
    check, files = checking_core_files()
    if check:
        layout_missing_core_files = [
            [sg.Image(
                r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\Exclamation mark.png'),
                sg.Text('Missing core files!\nYou must first finish setting up the core files!')]]
        missing_core_files_window = sg.Window('Error', layout_missing_core_files)
        pygame.mixer.music.load(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\sounds\error.wav')
        pygame.mixer.music.play(loops=0)
        missing_core_files_window.read()
        missing_core_files_window.close()
        return
    Update_boot_file_parameters_and_Lane_map_transfer()
    channel_list = []
    for i in range(number_of_lanes):
        for x in ['in', 'out', 'lane']:
            channel_list += ['channel_{i}_{x}'.format(i=i + 1, x=x)]

    while True:
        layout_windows_select = [
            [sg.Image(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\logo.png')],
            [sg.Text(font=("Arial", "2"))],
            [sg.Text('Please select channel', font=("Arial", "12"), size=(20, 2))],
            [[sg.Text('channel '), sg.Combo(channel_list, key='channel')]],
            [sg.Text(font=("Arial", "25"), size=4), sg.Button('start mask')]]
        select_window = sg.Window("Building mask", layout_windows_select)
        event, values = select_window.read()
        if event == sg.WIN_CLOSED:
            break
        if event == 'start mask':
            channel = values['channel']
            print(channel)
            select_window.close()
            image_capture(channel, mask_mode=True)
            layout_windows_paint = [
                [sg.Text(size=4),
                 sg.Image(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\logo.png')],
                [sg.Text(font=("Arial", "2"))],
                [sg.Text(' Paint in the color shown, the areas you want to mask', font=("Arial", "11"))],
                [sg.Text(size=48)],
                [sg.Text(size=14), sg.Text(' RGB = 181,229,29')],
                [[sg.Text(size=15),
                  sg.Image(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\mask color.png')]],
                [sg.Text('                              '), sg.Button('building mask')]]
            paint_window = sg.Window("Building mask", layout_windows_paint)
            while True:
                event, values = paint_window.read()
                if event == sg.WIN_CLOSED:
                    paint_window.close()
                    break
                if event == 'building mask':
                    paint_window.close()
                    image_mask = Image.open('Image mask.png')
                    pixelMap = image_mask.load()

                    mask = Image.new('RGB', (width, height))
                    pixelsNew = mask.load()
                    for i in range(mask.size[0]):
                        for j in range(mask.size[1]):
                            if pixelMap[i, j] == (181, 229, 29, 255) or pixelMap[i, j] == (181, 230, 29, 255) or\
                                    pixelMap[i, j] == (181, 229, 29) or pixelMap[i, j] == (181, 230, 29):
                                pixelsNew[i, j] = (0, 0, 0, 255)
                            else:
                                pixelsNew[i, j] = (255, 255, 255, 255)
                    image_mask.close()
                    os.remove(r'Image mask.png')
                    if os.path.exists(channels[channel][4]):
                        os.remove(channels[channel][4])
                    mask.save(channels[channel][4])
                    Mask_on_image(channel)
                    mask.close()

                    layout_windows_result = [
                        [sg.Text(size=14),
                         sg.Image(
                             r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\logo.png')],
                        [sg.Text(font=("Arial", "2"))],
                        [sg.Text(size=30), sg.Text('Result', font=("Arial", "12"))],
                        [sg.Text(size=12), sg.Text('mask', size=30), sg.Text('mask on image')],
                        [sg.Image(size=(300, 300), key='-mask-'),
                         sg.Text(),
                         sg.Image(size=(300, 300), key='-mask_on_image-')],
                        [sg.Text()], ]
                    result_window = sg.Window("Building mask", layout_windows_result, margins=(0, 0), finalize=True)
                    result_window['-mask-'].update(data=Resize_image_window((300, 300), channels[channel][4]))
                    result_window['-mask_on_image-'].update(
                        data=Resize_image_window((300, 300), channels[channel][2]))
                    while True:
                        event, values = result_window.read()
                        if event == sg.WIN_CLOSED:
                            result_window.close()
                            break

                    break


def Systems_matching():
    global number_of_lanes
    if not os.path.exists(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Boot file.xlsx'):
        layout_missing_boot_file = [
            [sg.Image(
                r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\Exclamation mark.png'),
                sg.Text('Missing boot file!\nPlease set up the boot file before setting up the Systems matching')]]
        missing_boot_file_window = sg.Window('Error', layout_missing_boot_file)
        pygame.mixer.music.load(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\sounds\error.wav')
        pygame.mixer.music.play(loops=0)
        missing_boot_file_window.read()
        missing_boot_file_window.close()
        return
    Update_boot_file_parameters_and_Lane_map_transfer()
    ser.write(b'$@SMS$')
    sensor_port = ''
    str_sensor_port = ''
    find_channel = []
    channel_list = []

    for i in range(number_of_lanes):
        for x in ['in', 'out', 'lane']:
            channel_list += ['channel_{i}_{x}'.format(i=i + 1, x=x)]
            find_channel += ['channel_{i}_{x}'.format(i=i + 1, x=x)]

    workbook_Systems_matching = Workbook()
    sheet_Systems_matching = workbook_Systems_matching.active
    sheet_Systems_matching["A1"] = "Channel"
    row = 2
    for channel_name in channel_list:
        sheet_Systems_matching["A{row}".format(row=row)] = channel_name
        row += 1
    sheet_Systems_matching["B1"] = "Camera"
    sheet_Systems_matching["C1"] = "Sensor port"
    camera_num = 1
    camera_num_new = 1

    layout_Systems_matching = [
        [sg.Text(size=30),
         sg.Image(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\logo.png')],
        [sg.Text()],
        [sg.Frame('Match camera, port and channel', [[sg.Text('', key='sensor port'), sg.Text('', key='port')],
                                                     [sg.Text('camera number '),
                                                      sg.Combo([i for i in range(1, number_camera + 1)],
                                                               camera_num, font="11", size=2, key='camera')],
                                                     [sg.Listbox(channel_list, size=(25, 19), key='channel')], ],
                  border_width=10),
         sg.Text(' '), sg.Image(filename='', key='image')],
        [sg.Text(size=(9, 2)), sg.Button('Match', size=7), sg.Text(size=76), sg.Button('Exit', size=7)]]
    window_Systems_matching = sg.Window('Systems matching', [layout_Systems_matching, ], location=(800, 400))
    Camera_resolution('Systems matching')

    while True:
        cap = globals()[f"camera{camera_num}"]
        while True:
            event, values = window_Systems_matching.Read(timeout=20, timeout_key='timeout'
                                                         )  # get events for the window with 20ms max wait
            data_input = data_transfer.read_command(ser)
            if event is None or event == 'Exit':  # if user closed window or press 'Exit', quit
                break
            if data_input != '' and data_input[0] + data_input[2] == 'pp':
                if data_input[3] == '0':
                    str_sensor_port = f'Port {data_input[1].upper()} Pin {data_input[4:]}'
                else:
                    str_sensor_port = f'Port {data_input[1].upper()} Pin {data_input[3:]}'
                sensor_port = data_input

            if len(values['channel']) != 0:
                if values['channel'][0][10:] != 'lane':
                    window_Systems_matching['sensor port'].update('sensor port: ')
                    window_Systems_matching['port'].update(str_sensor_port)
                else:
                    window_Systems_matching['sensor port'].update('')
                    window_Systems_matching['port'].update('')

            camera_num_new = values['camera']
            if camera_num_new != camera_num:
                break
            if event == 'Match' and len(values['channel']) != 0:
                if values['channel'][0][10:] != 'lane' and sensor_port != '':
                    sheet_Systems_matching[
                        "B{row}".format(row=find_channel.index(values['channel'][0]) + 2)] = camera_num
                    sheet_Systems_matching[
                        "C{row}".format(row=find_channel.index(values['channel'][0]) + 2)] = sensor_port
                    channel_list.remove(values['channel'][0])
                    window_Systems_matching['channel'].update(channel_list)
                elif values['channel'][0][10:] == 'lane':
                    sheet_Systems_matching[
                        "B{row}".format(row=find_channel.index(values['channel'][0]) + 2)] = camera_num
                    channel_list.remove(values['channel'][0])
                    window_Systems_matching['channel'].update(channel_list)
                window_Systems_matching['sensor port'].update('')
                window_Systems_matching['port'].update('')
                sensor_port = ''
                str_sensor_port = ''
                break
            window_Systems_matching.Element('image').Update(
                data=cv2.imencode('.png', cap.read()[1])[1].tobytes())  # Update image in window
        camera_num = camera_num_new
        if len(channel_list) == 0:
            workbook_Systems_matching.save(
                filename=
                r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Systems match file.xlsx')
            pygame.mixer.music.load(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\sounds\message.wav')
            pygame.mixer.music.play(loops=0)
            sg.popup('Data saved!')
            break
        elif event is None or event == 'Exit':
            layout_missing_boot_file = [
                [sg.Image(
                    r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\Exclamation mark.png')
                    , sg.Text('You have not finished all the matching so the file will not be saved!')]]
            missing_boot_file_window = sg.Window('Error!', layout_missing_boot_file)
            pygame.mixer.music.load(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\sounds\error.wav')
            pygame.mixer.music.play(loops=0)
            missing_boot_file_window.read()
            missing_boot_file_window.close()
            break
    window_Systems_matching.close()
    Camera_resolution()
    ser.write(b'$@SMF$')


def Lane_map():
    global number_of_lanes
    if not os.path.exists(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Boot file.xlsx'):
        layout_missing_boot_file = [
            [sg.Image(
                r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\Exclamation mark.png'),
                sg.Text('Missing boot file!\nPlease set up the boot file before setting up the lanes map')]]
        missing_boot_file_window = sg.Window('Error', layout_missing_boot_file)
        pygame.mixer.music.load(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\sounds\error.wav')
        pygame.mixer.music.play(loops=0)
        missing_boot_file_window.read()
        missing_boot_file_window.close()
        return
    Update_boot_file_parameters_and_Lane_map_transfer()
    layout_lane_map = [
        [sg.Text(size=17), sg.Image(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\logo.png')],
        [sg.Text()],
        [sg.Text(size=15), sg.Text('standby', size=14), sg.Text('Time in', size=13), sg.Text('Cars in', size=11),
         sg.Text('Enable delay')]]

    for i in range(number_of_lanes):
        layout_lane_map += [[sg.Text('Lane in {i}'.format(i=i + 1), size=15),
                             sg.Spin([i for i in range(1, 601)], font="11", size=3,
                                     key='standby lane {i}'.format(i=i + 1)),
                             sg.Text('sec', size=7),
                             sg.Spin([i for i in range(1, 601)], font="11", size=3,
                                     key='time lane {i}'.format(i=i + 1)),
                             sg.Text('sec', size=7),
                             sg.Spin([i for i in range(1, 100)], font="11", size=2,
                                     key='cars lane {i}'.format(i=i + 1)), sg.Text('car', size=7),
                             sg.Spin([i for i in range(1, 601)], font="11", size=3,
                                     key='enable delay {i}'.format(i=i + 1)),
                             sg.Text('sec', size=5)]]

    layout_lane_map += [[sg.Text()], [sg.Button('Save'), sg.Button('Exit')]]
    lane_map_window = sg.Window('Lane Map', layout_lane_map)
    lane_map_window.read(timeout=20)

    if os.path.exists(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Lane map.xlsx'):
        lane_file = pd.read_excel(
            r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Lane map.xlsx')
        for i in range(number_of_lanes):
            lane_map_window['standby lane {i}'.format(i=i + 1)].update(lane_file['standby'][i])
            lane_map_window['time lane {i}'.format(i=i + 1)].update(lane_file['Time in'][i])
            lane_map_window['cars lane {i}'.format(i=i + 1)].update(lane_file['Cars in'][i])
            lane_map_window['enable delay {i}'.format(i=i + 1)].update(lane_file['E delay'][i])
    while True:
        event, values = lane_map_window.read()
        if event == 'Save':
            pygame.mixer.music.load(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\sounds\message.wav')
            pygame.mixer.music.play(loops=0)
            sg.popup('Data saved!')
            workbook_lane_map = Workbook()
            sheet_lane_map = workbook_lane_map.active
            sheet_lane_map["B1"] = 'standby'
            sheet_lane_map["C1"] = 'Time in'
            sheet_lane_map["D1"] = 'Cars in'
            sheet_lane_map["E1"] = 'E delay'
            for i in range(number_of_lanes):
                sheet_lane_map["A{row}".format(row=i + 2)] = 'Lane in {i}'.format(i=i + 1)
                sheet_lane_map["B{row}".format(row=i + 2)] = values['standby lane {i}'.format(i=i + 1)]
                sheet_lane_map["C{row}".format(row=i + 2)] = values['time lane {i}'.format(i=i + 1)]
                sheet_lane_map["D{row}".format(row=i + 2)] = values['cars lane {i}'.format(i=i + 1)]
                sheet_lane_map["E{row}".format(row=i + 2)] = values['enable delay {i}'.format(i=i + 1)]
            workbook_lane_map.save(
                filename=r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Lane map.xlsx')
        lane_map_window.close()
        break


def System_boot():
    if os.path.exists(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Boot file.xlsx'):
        boot_file = pd.read_excel(
            r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Boot file.xlsx')

        layout_system_boot = [
            [sg.Text('   ', size=4, font=('', "2")),
             sg.Image(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\logo.png')],
            [sg.Text('')],
            [sg.Checkbox('Analyze', bool(boot_file['analyze'].tolist()[-1]), size=20, font="11", key='analyze'),
             sg.Text(' ', size=1, font=('', "4")), sg.Spin([i for i in range(2, 29)], boot_file['days'].tolist()[-1],
                                                           font="11", key='days', size=2),
             sg.Text('  days', font="11")],
            [sg.Text('')],
            [sg.Text('Lanes in the square', size=24, font="11"), sg.Spin([i for i in range(3, 10)],
                                                                         boot_file['lanes'].tolist()[-1], font="11",
                                                                         key='lanes'), sg.Text('    lane', font="11")],
            [sg.Text('Maximum cars in Level 2', size=24, font="11"), sg.Spin([i for i in range(10, 101)],
                                                                             boot_file['Level_2'].tolist()[-1],
                                                                             font="11", key='Level_2'),
             sg.Text('  cars', font="11")],
            [sg.Text('Maximum cars in Level 3', size=24, font="11"), sg.Spin([i for i in range(10, 101)],
                                                                             boot_file['Level_3'].tolist()[-1],
                                                                             font="11", key='Level_3'),
             sg.Text('cars', font="11")],
            [sg.Text('Minimum cars for statistics', size=24, font="11"), sg.Spin([i for i in range(5, 51)],
                                                                                 boot_file['min cars'].tolist()[-1],
                                                                                 font="11", key='min cars', size=2),
             sg.Text('  cars', font="11")],
            [sg.Text('Statistics time', size=24, font="11"), sg.Spin([i for i in range(1, 61)],
                                                                     boot_file['statistics'].tolist()[-1], font="11",
                                                                     key='statistics', size=2),
             sg.Text('  min', font="11")],
            [sg.Button('Save'), sg.Button('Exit')]]

    else:
        layout_system_boot = [
            [sg.Text('   ', size=4, font=('', "2")),
             sg.Image(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\logo.png')],
            [sg.Text('')],
            [sg.Checkbox('Analyze', size=20, font="11", key='analyze'), sg.Text(' ', size=1, font=('', "4")),
             sg.Spin([i for i in range(2, 29)], font="11", key='days', size=2), sg.Text('  days', font="11")],
            [sg.Text('')],
            [sg.Text('Lanes in the square', size=24, font="11"), sg.Spin([i for i in range(3, 10)],
                                                                         4, font="11", key='lanes'),
             sg.Text('    lane', font="11")],
            [sg.Text('Maximum cars in Level 2', size=24, font="11"), sg.Spin([i for i in range(10, 101)],
                                                                             24, font="11", key='Level_2'),
             sg.Text('  cars', font="11")],
            [sg.Text('Maximum cars in Level 3', size=24, font="11"), sg.Spin([i for i in range(10, 101)], 100,
                                                                             font="11", key='Level_3'),
             sg.Text('cars', font="11")],
            [sg.Text('Minimum cars for statistics', size=24, font="11"), sg.Spin([i for i in range(5, 51)],
                                                                                 5, font="11", key='min cars', size=2),
             sg.Text('  cars', font="11")],
            [sg.Text('Statistics time', size=24, font="11"), sg.Spin([i for i in range(1, 61)],
                                                                     7, font="11", key='statistics', size=2),
             sg.Text('  min', font="11")],
            [sg.Button('Save'), sg.Button('Exit')]]
    system_boot_window = sg.Window("System Boot", layout_system_boot)
    event, values = system_boot_window.read()
    if event == 'Save':
        pygame.mixer.music.load(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\sounds\message.wav')
        pygame.mixer.music.play(loops=0)
        sg.popup('Data saved!')
        if not os.path.exists(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Boot file.xlsx'):
            workbook = Workbook()
            workbook.save(filename=r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Boot file.xlsx')
        boot_file = pd.read_excel(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Boot file.xlsx')
        boot_file = boot_file.append(values, ignore_index=True)
        boot_file.to_excel(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Boot file.xlsx',
                           index=False)
    system_boot_window.close()


def Channel_test():
    global channel_test
    channel_test = True
    check, files = checking_core_files()
    if check:
        layout_missing_core_files = [
            [sg.Image(
                r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\Exclamation mark.png'),
                sg.Text('Missing core files!\nYou must first finish setting up the core files!')]]
        missing_core_files_window = sg.Window('Error', layout_missing_core_files)
        pygame.mixer.music.load(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\sounds\error.wav')
        pygame.mixer.music.play(loops=0)
        missing_core_files_window.read()
        missing_core_files_window.close()
        return
    Update_boot_file_parameters_and_Lane_map_transfer()
    channel_list = []
    for i in range(number_of_lanes):
        for x in ['in', 'out', 'lane']:
            channel_list += ['channel_{i}_{x}'.format(i=i + 1, x=x)]
    layout_Channel_test = [
        [sg.Image(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\logo.png')],
        [sg.Text('')],
        [sg.Text('select channel ', size=15, font="11"), sg.Combo(channel_list, size=15, key='channel')],
        [sg.Text('', size=15, font=('', "1"))],
        [sg.Checkbox('with draw boxes', size=15, font="11", key='draw_boxes')],
        [sg.Text('', size=15, font=('', "1"))],
        [sg.Checkbox('without camera', size=15, font="11", key='without_camera')],
        [sg.Text()],
        [sg.Text(size=12), sg.Button(' Run '), sg.Button(' Exit ')]]
    Channel_test_window = sg.Window('Channel test', layout_Channel_test)
    while True:
        event, values = Channel_test_window.read()
        if event == ' Run ' and len(values['channel']) > 0:
            print(values['channel'])
            channel = values['channel']
            draw_boxes = values['draw_boxes']
            without_camera = values['without_camera']
            image_with_bounding_box = detect_img(channel, draw_boxes=draw_boxes, without_camera=without_camera)
            if draw_boxes and image_with_bounding_box is not None:
                image_with_bounding_box.save(r'@#$%^&@%%.jpg')
                layout_windows_result = [
                    [sg.Text(size=17),
                     sg.Image(
                         r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\logo.png')],
                    [sg.Text(font=("Arial", "2"))],
                    [sg.Text(size=32), sg.Text('Result', font=("Arial", "12"))],
                    [sg.Text(size=12)],
                    [sg.Text(), sg.Image(size=(648, 486), key='-result-'), sg.Text()],
                    [sg.Text('', size=15, font=('', "1"))], ]
                result_window = sg.Window("Channel test", layout_windows_result, margins=(0, 0), finalize=True)
                result_window['-result-'].update(data=Resize_image_window((648, 486), r'@#$%^&@%%.jpg'))
                while True:
                    event2, values2 = result_window.read()
                    if event2 == sg.WIN_CLOSED:
                        result_window.close()
                        os.remove(r'@#$%^&@%%.jpg')
                        break
            print('Run')
        if event == sg.WIN_CLOSED or event == ' Exit ':
            channel_test = False
            Channel_test_window.close()
            break


def administrator():
    global number_of_lanes
    while True:
        Update_boot_file_parameters_and_Lane_map_transfer()
        layout_administrator = [
            [sg.Text(size=1),
             sg.Image(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\logo.png')],
            [sg.Text('')],
            [sg.Frame('Core files', [
                [sg.Button('System boot'), sg.Button('Systems matching'), sg.Button('Lane map')]
            ], border_width=5)],
            [sg.Text()],
            [sg.Text(size=1), sg.Button('Building mask'), sg.Button('Channel test'), sg.Button('System test')],
            [sg.Text()],
            [sg.Text(), sg.Button('Exit')]]
        administrator_window = sg.Window("Admin Mode", layout_administrator)
        event, values = administrator_window.read()
        if event == sg.WIN_CLOSED or event == 'Exit':
            administrator_window.close()
            check, files = checking_core_files()
            if check:
                goodbye()
            else:
                layout_missing_boot_file = [
                    [sg.Image(
                        r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\Exclamation mark.png')
                        , sg.Text('It is recommended to back up all the "Core files" folders!')]]
                missing_boot_file_window = sg.Window('Message', layout_missing_boot_file)
                pygame.mixer.music.load(
                    r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\sounds\message.wav')
                pygame.mixer.music.play(loops=0)
                missing_boot_file_window.read(timeout=5000)
                missing_boot_file_window.close()
            break
        if event == 'System boot':
            Update_boot_file_parameters_and_Lane_map_transfer()
            administrator_window.close()
            System_boot()
        if event == 'Systems matching':
            administrator_window.close()
            Systems_matching()
        if event == 'Lane map':
            administrator_window.close()
            Lane_map()
        if event == 'Building mask':
            administrator_window.close()
            Building_mask()
        if event == 'Channel test':
            administrator_window.close()
            Channel_test()
        if event == 'System test':
            administrator_window.close()
            System_test()


def administrator_sign_in():
    layout_sign_in = [
        [sg.Text("Please enter your username and password", font=40)],
        [sg.Text("Username", font=10), sg.InputText(key='-usrnm-', font=16)],
        [sg.Text("Password", font=10), sg.InputText(key='-pwd-', password_char='*', font=16)],
        [sg.Text('', k='try')],
        [sg.Button('Sign in'), sg.Button('Exit')]]
    layout_last_try = [
        [sg.Image(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\Exclamation mark.png'),
         sg.Text('This is your last try!\nAfter that, the program will close!!!!')]]
    layout_Welcome = [[sg.Text("     Welcome administrator", font=40, size=(25, 2))]]

    last_try_window = sg.Window('administrator', layout_last_try)
    sign_in_window = sg.Window('administrator', layout_sign_in)

    Counted_attempts = 0
    while True:
        event, values = sign_in_window.read()
        if event == sg.WIN_CLOSED or event == 'Exit':
            sign_in_window.close()
            break
        if event == 'Sign in':
            if values['-usrnm-'] in administrator_dict:
                if values['-pwd-'] == administrator_dict[values['-usrnm-']]:
                    sign_in_window.close()
                    check, files = checking_core_files()
                    if check:
                        layout_Welcome = [[sg.Text("     Welcome administrator", font=40, size=(25, 1))],
                                          [sg.Text('Missing files:\n   *' + '\n   *'.join(files))]]
                    Welcome_window = sg.Window('administrator', layout_Welcome)
                    pygame.mixer.music.load(
                        r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\sounds\message.wav')
                    pygame.mixer.music.play(loops=0)
                    Welcome_window.read(timeout=5000)
                    Welcome_window.close()
                    administrator()
                    break

            if values['-usrnm-'] != '' and values['-pwd-'] != '':
                Counted_attempts += 1
                sign_in_window['try'].update('The username or password is incorrect\n'
                                             f'You have {3 - Counted_attempts} attempts left')
                if Counted_attempts > 3:
                    sign_in_window.close()
                    goodbye()
                elif Counted_attempts > 2:
                    pygame.mixer.music.load(
                        r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\sounds\error.wav')
                    pygame.mixer.music.play(loops=0)
                    last_try_window.read(timeout=2000)
                    last_try_window.close()


check, files = checking_core_files()
if check:
    layout_core_files_error = [
        [sg.Image(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\Exclamation mark.png'),
         sg.Text('The system needs to be booted!!\nPlease call the system administrator !!!!')],
        [sg.Text(' ', size=18), sg.Button('OK')]]
    core_files_error_window = sg.Window('administrator', layout_core_files_error)
    pygame.mixer.music.load(
        r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\sounds\error.wav')
    pygame.mixer.music.play(loops=0)
    event, values = core_files_error_window.read()
    if event == sg.WIN_CLOSED:
        core_files_error_window.close()
        goodbye()
    if event == 'OK':
        core_files_error_window.close()
        administrator_sign_in()
check, files = checking_core_files()
if check:
    goodbye()
sg.theme('DarkAmber')
layout_windows_start = [
    [sg.Image(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\logos\logo start.png')],
    [sg.Text(" " * 2, size=(45, 2)), sg.Button('administrator', size=(15, 2))]]

pygame.mixer.music.load(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\windows\sounds\start.wav')
pygame.mixer.music.play(loops=0)
window = sg.Window("Welcome", layout_windows_start)
event, values = window.read(timeout=8000)
if event == 'administrator':
    pygame.mixer.music.stop()
    window.close()
    administrator_sign_in()
if event == sg.WIN_CLOSED:
    pygame.mixer.music.stop()
window.close()


match_file = pd.read_excel(r'venv\Lib\site-packages\Yakir_and_Kfir_Project\Files\Core files\Systems match file.xlsx')


def main():
    global clock_enabling
    global in_process
    global checks_counter
    global problem_lane
    global crowded_square

    Update_boot_file_parameters_and_Lane_map_transfer()
    port_dict = {}
    index_channel = 0
    checks_counter = 0
    print('main')

    for i in match_file['Sensor port']:
        if i != 'NaN':
            port_dict[i] = match_file['Channel'][index_channel]
        index_channel += 1

    clock_enabling = True
    interrupt_time = threading.Timer(1.0, Int_time)
    interrupt_time.start()

    while True:
        data = data_transfer.read_command(ser)
        if data != '':
            if not crowded_square:
                if data in port_dict.keys():
                    clock_enabling = False
                    if 'in' in port_dict[data]:
                        if channels[port_dict[data]][6] != '':
                            enter_to_square(port_dict[data])
                        else:
                            print(port_dict[data] + ' capture')
                            detect_img(port_dict[data])

                    elif 'out' in port_dict[data]:
                        if not channels[port_dict[data]][11]:
                            channels[port_dict[data]][11] = True
                            print(port_dict[data] + ' front')
                        else:
                            print(port_dict[data] + ' capture')
                            detect_img(port_dict[data])
                            channels[port_dict[data]][11] = False
                            if channels[port_dict[data]][6] == '':
                                if channels[port_dict[data]][12] >= 2:
                                    channels[port_dict[data]][12] = 0
                                    channels[port_dict[data]][11] = True
                                else:
                                    channels[port_dict[data]][12] += 1

                    clock_enabling = True
                    interrupt_time = threading.Timer(1.0, Int_time)
                    interrupt_time.start()
                    if not('out' in port_dict[data] and channels[port_dict[data]][11]):
                        print('level_2:')
                        for v in level_2.items():
                            print("%s -> %s" % v)
                        print('level_3:')
                        for v in level_3.items():
                            print("%s -> %s" % v)

            if 'ut ' in data:
                tables_fun.creating_problem_document_UT(data, time_document)

            elif 'pt ' in data:
                tables_fun.creating_problem_document_PT(data, time_document, system_test)

            elif data == 'CHECK':
                Check_transmission()
                checks_counter += 1

            elif 'RT ' in data:
                tables_fun.creating_problem_document_RT(data, time_document, system_test)
                print([data])

            elif 'SOLVE = ' in data:
                checks_counter = 0
                if not crowded_square:
                    tables_fun.creating_problem_document_SOLVE(data, time_document, system_test)
                else:
                    solve = data[:5] + " = Crowded square"
                    tables_fun.creating_problem_document_SOLVE(solve, time_document, system_test)
                tables_fun.creating_problem_document_end(level_1, level_2, level_3, problem_license_plate,
                                                         time_document, system_test)
                print([data])
                in_process = False


if __name__ == "__main__":
    main()