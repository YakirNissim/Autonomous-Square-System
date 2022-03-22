from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import pathlib
import time
import random
import os
from datetime import timedelta

# Font text??????
font_levels = Font(bold=True, color="8a2be2", size=13)
font_titles = Font(bold=True, color="0000ff", size=12)
text_centering = Alignment(horizontal="center")
bold_red_text_font = Font(bold=True, color="ff0000")
columns = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
           'U', 'V', 'W', 'X', 'Y', 'Z']
titles = ["license plate", 'status', 'Standby time', 'lane in', 'lane out', 'time in', 'time out']
clean_vector = ['-', '0', '00:00', '0', '0', '00:00:00', '00:00:00']


def manual_file(number_of_lanes=4):
    workbook = Workbook()
    sheet = workbook.active

    sheet["A1"].font = font_levels
    sheet["A1"] = "Level 1"
    for i in range(number_of_lanes):
        sheet["{column}1".format(column=columns[i])].font = font_titles
        sheet["{column}1".format(column=columns[i])] = "lane {num}".format(num=i + 1)
        sheet["{column}2".format(column=columns[i])].alignment = text_centering
        sheet["{column}2".format(column=columns[i])] = '0'

    sheet["A4"].font = font_levels
    sheet["A4"] = "Level 2"
    for i in range(7):
        sheet["{column}4".format(column=columns[i])].font = font_titles
        sheet["{column}4".format(column=columns[i])] = titles[i]
        sheet["{column}5".format(column=columns[i])].alignment = text_centering
        sheet["{column}5".format(column=columns[i])] = clean_vector[i]

    sheet["A7"].font = font_levels
    sheet["A7"] = "Level 3"
    for i in range(7):
        sheet["{column}7".format(column=columns[i])].font = font_titles
        sheet["{column}7".format(column=columns[i])] = titles[i]
        sheet["{column}8".format(column=columns[i])].alignment = text_centering
        sheet["{column}8".format(column=columns[i])] = clean_vector[i]

    workbook.save(filename=r"Files\System test\Manual file.xlsx")


def manual_operation(number_of_lanes=4):
    level_1 = []
    level_2 = {}
    level_3 = {}
    if not os.path.exists(r"Files\System test\Manual file.xlsx"):
        return level_1, level_2, level_3
    workbook = load_workbook(filename=r"Files\System test\Manual file.xlsx")
    workbook.sheetnames
    sheet = workbook.active

    pointer_row = 1
    key = ''
    value = []

    # level 1
    while sheet["A{row}".format(row=pointer_row)].value != 'Level 1':
        pointer_row += 1
        if pointer_row > 100:
            print('error: The file is not written as required!!!')
            quit()

    for i in range(number_of_lanes):
        level_1 += [str(sheet["{column}{row}".format(column=columns[i], row=pointer_row + 1)].value)]

    # level 2
    while sheet["A{row}".format(row=pointer_row)].value != 'Level 2':
        pointer_row += 1
        if pointer_row > 200:
            print('error: The file is not written as required!!!')
            quit()
    pointer_row += 1

    while sheet["A{row}".format(row=pointer_row + 1)].value != 'Level 3' and \
            sheet["B{row}".format(row=pointer_row)].value is not None:

        if str(sheet["B{row}".format(row=pointer_row)].value) == '-':
            key = str(random.randint(11111111, 99999999))
            while key in level_2.keys():
                key = str(random.randint(11111111, 99999999))
        else:
            key = str(sheet["B{row}".format(row=pointer_row)].value)
        value = []
        for i in range(6):
            value += [str(sheet["{column}{row}".format(column=columns[i + 1], row=pointer_row)].value)]
        time_now = time.strftime('%M:%S')
        time_now_convert = timedelta(minutes=int(time_now[0:2]), seconds=int(time_now[3:]))
        if len(value[1]) == 8:
            standby_convert = timedelta(minutes=int(value[1][0:2]), seconds=int(value[1][3:5]))
        else:
            standby_convert = timedelta(minutes=int(value[1][0:2]), seconds=int(value[1][3:]))
        value += [str(time_now_convert - standby_convert)[2:]]
        level_2[key] = value
        pointer_row += 1

    # level 3
    while sheet["A{row}".format(row=pointer_row)].value != 'Level 3':
        pointer_row += 1
        if pointer_row > 200:
            print('error: The file is not written as required!!!')
            quit()
    pointer_row += 1

    while sheet["B{row}".format(row=pointer_row)].value is not None:
        if str(sheet["B{row}".format(row=pointer_row)].value) == '-':
            key = str(random.randint(11111111, 99999999))
            while key in level_3.keys() and key in level_2.keys():
                key = str(random.randint(11111111, 99999999))
        else:
            key = str(sheet["B{row}".format(row=pointer_row)].value)
        value = []
        for i in range(6):
            value += [str(sheet["{column}{row}".format(column=columns[i + 1], row=pointer_row)].value)]
        time_now = time.strftime('%M:%S')
        time_now_convert = timedelta(minutes=int(time_now[0:2]), seconds=int(time_now[3:]))
        standby_convert = timedelta(minutes=int(value[1][0:2]), seconds=int(value[1][3:]))
        value += [str(time_now_convert - standby_convert)[2:]]
        level_3[key] = value
        pointer_row += 1

    return level_1, level_2, level_3


def creating_problem_document_start(level_1, level_2, level_3, problem_license_plate, time_document, system_test=False):
    date = time.strftime("%d.%m.%Y")
    workbook = Workbook()
    sheet = workbook.active

    if system_test:
        if not (os.path.exists(r"Files\System test\{date}".format(date=date))):
            os.makedirs(r"Files\System test\{date}".format(date=date))
    else:
        if not (os.path.exists(r"Files\Problem report\{date}".format(date=date))):
            os.makedirs(r"Files\Problem report\{date}".format(date=date))

    size_1 = len(level_1)
    sheet["A1"].font = font_levels
    sheet["A1"] = "Level 1"
    for i in range(size_1):
        sheet["{column}1".format(column=columns[i])].font = font_titles
        sheet["{column}1".format(column=columns[i])] = "lane {num}".format(num=i + 1)
        sheet["{column}2".format(column=columns[i])].alignment = text_centering
        sheet["{column}2".format(column=columns[i])] = level_1[i]

    sheet["A4"].font = font_levels
    sheet["A4"] = "Level 2"
    for i in range(7):
        sheet["{column}4".format(column=columns[i])].font = font_titles
        sheet["{column}4".format(column=columns[i])] = titles[i]
    vector = 0
    for key in level_2.keys():
        if key == problem_license_plate:
            sheet["B{row}".format(row=5 + vector)].font = bold_red_text_font
            sheet["D{row}".format(row=5 + vector)].font = bold_red_text_font
        sheet["B{row}".format(row=5 + vector)].alignment = text_centering
        sheet["B{row}".format(row=5 + vector)] = key
        for i in range(6):
            sheet["{column}{row}".format(column=columns[i + 1], row=5 + vector)].alignment = text_centering
            sheet["{column}{row}".format(column=columns[i + 1], row=5 + vector)] = level_2[key][i]
        vector += 1

    sheet["A{row}".format(row=6 + vector)].font = font_levels
    sheet["A{row}".format(row=6 + vector)] = "Level 3"
    for i in range(7):
        sheet["{column}{row}".format(column=columns[i], row=6 + vector)].font = font_titles
        sheet["{column}{row}".format(column=columns[i], row=6 + vector)] = titles[i]
    for key in level_3.keys():
        sheet["B{row}".format(row=7 + vector)].alignment = text_centering
        sheet["B{row}".format(row=7 + vector)] = key
        for i in range(6):
            sheet["{column}{row}".format(column=columns[i + 1], row=7 + vector)].alignment = text_centering
            sheet["{column}{row}".format(column=columns[i + 1], row=7 + vector)] = level_3[key][i]
        vector += 1

    if system_test:
        workbook.save(filename=r"Files\System test\{date}\{name}.xlsx".format(
            date=date, name=time_document))
    else:
        workbook.save(filename=r"Files\Problem report\{date}\{name}.xlsx".format(
            date=date, name=time_document))


def creating_problem_document_PT(data, time_document, system_test=False):
    date = time.strftime("%d.%m.%Y")
    if system_test:
        workbook = load_workbook(filename=r"Files\System test\{date}\{name}.xlsx".format(
            date=date, name=time_document))
    else:
        workbook = load_workbook(filename=r"Files\Problem report\{date}\{name}.xlsx".format(
            date=date, name=time_document))
    sheet = workbook.active
    data = data[3:]
    data = data.split('  //')
    data_list_temp1 = []
    data_list_temp2 = []
    PT = []
    for i in range(4):
        data_list_temp1 += [data[i].split('  ')]

    for i in range(4):
        data_list_temp2 += [[data_list_temp1[i][0].split(' ')] + [data_list_temp1[i][1].split(' ')] +
                            [data_list_temp1[i][2].split(' ')] + [data_list_temp1[i][3].split(' ')]]
    for i in range(4):
        PT += [[[int(data_list_temp2[i][0][0]), bin(int(data_list_temp2[i][0][1]))[2:].zfill(4)],
                [int(data_list_temp2[i][1][0]), bin(int(data_list_temp2[i][1][1]))[2:].zfill(4)],
                [int(data_list_temp2[i][2][0]), bin(int(data_list_temp2[i][2][1]))[2:].zfill(4)],
                [int(data_list_temp2[i][3][0]), bin(int(data_list_temp2[i][3][1]))[2:].zfill(4)]]]
    sheet["L1"].alignment = text_centering
    sheet["L1"] = "PT"
    sheet["J2"].alignment = text_centering
    sheet["J2"] = "lane in"
    for i in range(4):
        sheet[f"J{3+i}"] = f"lane in {i+1}"
    for i in range(4):
        sheet[f"{columns[9+i]}2"].alignment = text_centering
        sheet[f"{columns[9+i]}2"] = f"lane out {i+1}"
    for row in range(4):
        for col in range(4):
            sheet[f"{columns[col + 9]}{3 + row}"].alignment = text_centering
            sheet[f"{columns[col+9]}{3 + row}"] = str(PT[row][col][0]) + ', ' + PT[row][col][1]

    if system_test:
        workbook.save(filename=r"Files\System test\{date}\{name}.xlsx".format(
            date=date, name=time_document))
    else:
        workbook.save(filename=r"Files\Problem report\{date}\{name}.xlsx".format(
            date=date, name=time_document))
    print('PT:')
    print(PT[0])
    print(PT[1])
    print(PT[2])
    print(PT[3])


def creating_problem_document_UT(data, time_document, system_test=False):
    date = time.strftime("%d.%m.%Y")
    if system_test:
        workbook = load_workbook(filename=r"Files\System test\{date}\{name}.xlsx".format(
            date=date, name=time_document))
    else:
        workbook = load_workbook(filename=r"Files\Problem report\{date}\{name}.xlsx".format(
            date=date, name=time_document))
    sheet = workbook.active
    data = data[3:-1]
    data = data.split(' ')
    UT = []
    for i in range(4):
        UT += [int(data[i])]
    print('UT:')
    print(UT)
    sheet["L8"].alignment = text_centering
    sheet["L8"] = "UT"
    sheet[f"{columns[8]}10"].alignment = text_centering
    sheet[f"{columns[8]}10"] = 'Using'
    for i in range(4):
        sheet[f"{columns[9+i]}9"].alignment = text_centering
        sheet[f"{columns[9+i]}9"] = f"lane in {i+1}"
        sheet[f"{columns[9+i]}10"].alignment = text_centering
        sheet[f"{columns[9+i]}10"] = str(UT[i]) + '%'

    if system_test:
        workbook.save(filename=r"Files\System test\{date}\{name}.xlsx".format(
            date=date, name=time_document))
    else:
        workbook.save(filename=r"Files\Problem report\{date}\{name}.xlsx".format(
            date=date, name=time_document))


def creating_problem_document_RT(data, time_document, system_test=False):
    date = time.strftime("%d.%m.%Y")
    if system_test:
        workbook = load_workbook(filename=r"Files\System test\{date}\{name}.xlsx".format(
            date=date, name=time_document))
    else:
        workbook = load_workbook(filename=r"Files\Problem report\{date}\{name}.xlsx".format(
            date=date, name=time_document))
    sheet = workbook.active
    data = data[3:-2]
    data = data.split(' /')
    RT = []
    data_list_temp1 = []
    for i in range(3):
        data_list_temp1 += [data[i].split(' ')]
    for i in range(3):
        RT += [[float(data_list_temp1[i][0]), int(data_list_temp1[i][1])]]
    sheet["L12"].alignment = text_centering
    sheet["L12"] = "RT"
    for col in range(3):
        sheet[f"{columns[col + 9]}13"].alignment = text_centering
        sheet[f"{columns[col + 9]}13"] = str(RT[col][0]) + ',  ' + str(RT[col][1])

    if system_test:
        workbook.save(filename=r"Files\System test\{date}\{name}.xlsx".format(
            date=date, name=time_document))
    else:
        workbook.save(filename=r"Files\Problem report\{date}\{name}.xlsx".format(
            date=date, name=time_document))


def creating_problem_document_SOLVE(data, time_document, system_test=False):
    date = time.strftime("%d.%m.%Y")
    if system_test:
        workbook = load_workbook(filename=r"Files\System test\{date}\{name}.xlsx".format(
            date=date, name=time_document))
    else:
        workbook = load_workbook(filename=r"Files\Problem report\{date}\{name}.xlsx".format(
            date=date, name=time_document))
    sheet = workbook.active
    sheet["L15"].alignment = text_centering
    sheet["L15"] = data[:5]
    sheet["L16"].alignment = text_centering
    sheet["L16"] = data[8:]
    if system_test:
        workbook.save(filename=r"Files\System test\{date}\{name}.xlsx".format(
            date=date, name=time_document))
    else:
        workbook.save(filename=r"Files\Problem report\{date}\{name}.xlsx".format(
            date=date, name=time_document))


def creating_problem_document_CHECK(check, number_check, time_document, system_test=False):
    date = time.strftime("%d.%m.%Y")
    if system_test:
        workbook = load_workbook(filename=r"Files\System test\{date}\{name}.xlsx".format(
            date=date, name=time_document))
    else:
        workbook = load_workbook(filename=r"Files\Problem report\{date}\{name}.xlsx".format(
            date=date, name=time_document))
    sheet = workbook.active

    sheet["L18"].alignment = text_centering
    sheet["L18"] = 'Cars wait'
    sheet["M18"].alignment = text_centering
    sheet["M18"] = 'Cars in'
    if number_check == 1:
        sheet["K19"].alignment = text_centering
        sheet["K19"] = f'CHECK {number_check}'
        sheet["L19"].alignment = text_centering
        sheet["L19"] = check[0]
        sheet["M19"].alignment = text_centering
        sheet["M19"] = check[1]
    elif number_check == 2:
        sheet["K20"].alignment = text_centering
        sheet["K20"] = f'CHECK {number_check}'
        sheet["L20"].alignment = text_centering
        sheet["L20"] = check[0]
        sheet["M20"].alignment = text_centering
        sheet["M20"] = check[1]
    elif number_check == 3:
        sheet["K21"].alignment = text_centering
        sheet["K21"] = f'CHECK {number_check}'
        sheet["L21"].alignment = text_centering
        sheet["L21"] = check[0]
        sheet["M21"].alignment = text_centering
        sheet["M21"] = check[1]
    if system_test:
        workbook.save(filename=r"Files\System test\{date}\{name}.xlsx".format(
            date=date, name=time_document))
    else:
        workbook.save(filename=r"Files\Problem report\{date}\{name}.xlsx".format(
            date=date, name=time_document))


def creating_problem_document_end(level_1, level_2, level_3, problem_license_plate, time_document, system_test=False):
    date = time.strftime("%d.%m.%Y")
    if system_test:
        workbook = load_workbook(filename=r"Files\System test\{date}\{name}.xlsx".format(
            date=date, name=time_document))
    else:
        workbook = load_workbook(filename=r"Files\Problem report\{date}\{name}.xlsx".format(
            date=date, name=time_document))
    sheet = workbook.active

    size_1 = len(level_1)
    sheet["P1"].font = font_levels
    sheet["P1"] = "Level 1"
    for i in range(size_1):
        sheet["{column}1".format(column=columns[i+15])].font = font_titles
        sheet["{column}1".format(column=columns[i+15])] = "lane {num}".format(num=i + 1)
        sheet["{column}2".format(column=columns[i+15])].alignment = text_centering
        sheet["{column}2".format(column=columns[i+15])] = level_1[i]

    sheet["P4"].font = font_levels
    sheet["P4"] = "Level 2"
    for i in range(7):
        sheet["{column}4".format(column=columns[i+15])].font = font_titles
        sheet["{column}4".format(column=columns[i+15])] = titles[i]
    vector = 0
    for key in level_2.keys():
        if key == problem_license_plate:
            sheet["Q{row}".format(row=5 + vector)].font = bold_red_text_font
            sheet["S{row}".format(row=5 + vector)].font = bold_red_text_font
        sheet["Q{row}".format(row=5 + vector)].alignment = text_centering
        sheet["Q{row}".format(row=5 + vector)] = key
        for i in range(6):
            sheet["{column}{row}".format(column=columns[i + 16], row=5 + vector)].alignment = text_centering
            sheet["{column}{row}".format(column=columns[i + 16], row=5 + vector)] = level_2[key][i]
        vector += 1

    sheet["P{row}".format(row=6 + vector)].font = font_levels
    sheet["P{row}".format(row=6 + vector)] = "Level 3"
    for i in range(7):
        sheet["{column}{row}".format(column=columns[i+15], row=6 + vector)].font = font_titles
        sheet["{column}{row}".format(column=columns[i+15], row=6 + vector)] = titles[i]
    for key in level_3.keys():
        sheet["Q{row}".format(row=7 + vector)].alignment = text_centering
        sheet["Q{row}".format(row=7 + vector)] = key
        for i in range(6):
            sheet["{column}{row}".format(column=columns[i + 16], row=7 + vector)].alignment = text_centering
            sheet["{column}{row}".format(column=columns[i + 16], row=7 + vector)] = level_3[key][i]
        vector += 1

    if system_test:
        workbook.save(filename=r"Files\System test\{date}\{name}.xlsx".format(
            date=date, name=time_document))
    else:
        workbook.save(filename=r"Files\Problem report\{date}\{name}.xlsx".format(
            date=date, name=time_document))